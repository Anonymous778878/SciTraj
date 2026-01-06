#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
evobench_ml_toolkit.py (UPDATED FOR NEW JSON FORMAT)

End-to-end toolkit for EvoBench-ML:
  - build: create dataset (papers/events/edges + splits) from JSON input files
  - validatepp: automated validation++ (structural, triangulation, shuffled-time control, bootstrap CIs)
  - review_pack: export CSV+XLSX for fast human audit (extractions + edges)
  - llm_judge: open-source LLM as zero-shot judge (extraction type + edge plausibility), with caching
  - stats: compute agreement (Cohen's kappa) and significance (bootstrap tests, CI)

INPUT FORMAT CHANGES:
  - all_sample.json: Contains query results with papers in nested 'data' array
  - imrad_corpus.json: Contains fulltext sections (Introduction, Methods, Results, Discussion, Conclusion)

Dependencies:
  - numpy
  - openpyxl
  - transformers, torch (optional, for LLM judge)
"""

import os, re, json, math, argparse, hashlib, random, time
from dataclasses import dataclass
from typing import Dict, Any, List, Tuple, Optional, Set
from collections import defaultdict, Counter

import numpy as np

# -----------------------
# Optional LLM imports (lazy)
# -----------------------
_TRANSFORMERS_OK = False
try:
    import torch
    from transformers import AutoTokenizer, AutoModelForCausalLM
    _TRANSFORMERS_OK = True
except Exception:
    _TRANSFORMERS_OK = False

# -----------------------
# XLSX export
# -----------------------
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================
# Utilities
# ============================================================

def ensure_dir(path: str) -> None:
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)

def read_jsonl(path: str):
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                yield json.loads(line)

def write_jsonl(path: str, rows: List[Dict[str, Any]]) -> None:
    ensure_dir(os.path.dirname(path))
    with open(path, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

def normalize_text(t: str) -> str:
    t = (t or "").lower()
    t = re.sub(r"\s+", " ", t).strip()
    return t

def tokenize(text: str) -> List[str]:
    text = normalize_text(text)
    return re.findall(r"[a-z0-9]+(?:[-/][a-z0-9]+)?", text)

def stable_hash_int(s: str) -> int:
    h = hashlib.sha256((s or "").encode("utf-8")).hexdigest()
    return int(h[:16], 16)

def clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))

def split_sentences(text: str) -> List[str]:
    text = re.sub(r"\s+", " ", (text or "")).strip()
    if not text:
        return []
    parts = re.split(r"(?<=[\.\?\!])\s+", text)
    return [p.strip() for p in parts if p.strip()]

def dedup_keep_order(xs: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in xs:
        k = normalize_text(x)
        if k and k not in seen:
            out.append(x)
            seen.add(k)
    return out

def jaccard(a: List[str], b: List[str]) -> float:
    sa, sb = set(a), set(b)
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)

def wilson_ci(k: int, n: int, z: float = 1.96) -> Tuple[float, float]:
    if n <= 0:
        return (0.0, 0.0)
    phat = k / n
    denom = 1 + z*z/n
    center = (phat + z*z/(2*n)) / denom
    margin = (z * math.sqrt((phat*(1-phat) + z*z/(4*n)) / n)) / denom
    return (max(0.0, center - margin), min(1.0, center + margin))

def paired_bootstrap_test(a: List[float], b: List[float], iters: int = 5000, seed: int = 13) -> Dict[str, Any]:
    """
    Paired bootstrap over per-example scores.
    Returns CI for mean difference and p-value for diff <= 0.
    """
    assert len(a) == len(b)
    rng = np.random.default_rng(seed)
    n = len(a)
    diffs = []
    a = np.array(a, dtype=np.float32)
    b = np.array(b, dtype=np.float32)
    for _ in range(iters):
        idx = rng.integers(0, n, size=n)
        diffs.append(float(np.mean(a[idx] - b[idx])))
    diffs = np.array(diffs)
    lo, hi = np.percentile(diffs, [2.5, 97.5])
    p = float(np.mean(diffs <= 0.0))
    return {"mean_diff": float(np.mean(a - b)), "ci95": [float(lo), float(hi)], "p_value": p}

def cohens_kappa(labels_a: List[str], labels_b: List[str]) -> float:
    """
    Cohen's kappa for categorical labels (same length).
    """
    assert len(labels_a) == len(labels_b)
    n = len(labels_a)
    if n == 0:
        return 0.0
    cats = sorted(set(labels_a) | set(labels_b))
    idx = {c:i for i,c in enumerate(cats)}
    cm = np.zeros((len(cats), len(cats)), dtype=np.int32)
    for x, y in zip(labels_a, labels_b):
        cm[idx[x], idx[y]] += 1
    po = np.trace(cm) / n
    px = cm.sum(axis=1) / n
    py = cm.sum(axis=0) / n
    pe = float(np.sum(px * py))
    if abs(1 - pe) < 1e-9:
        return 0.0
    return float((po - pe) / (1 - pe))


# ============================================================
# Expanded lexicons (methods/tasks/datasets/metrics/cues)
# ============================================================

METHOD_PATTERNS = {
    "transformer": [r"\btransformer(s)?\b", r"\bself-attention\b"],
    "bert_family": [r"\bbert\b", r"\broberta\b", r"\bdeberta\b", r"\belectra\b", r"\balbert\b"],
    "gpt_family": [r"\bgpt[- ]?[0-9]*\b", r"\bchatgpt\b", r"\binstructgpt\b", r"\bgpt-?neo\b"],
    "t5_family": [r"\bt5\b", r"\bmt5\b", r"\bbyt5\b", r"\bflan\b"],
    "seq2seq": [r"\bseq2seq\b", r"\bsequence[- ]to[- ]sequence\b", r"\bencoder[- ]decoder\b"],
    "prompting": [r"\bprompt(ing)?\b", r"\bfew[- ]shot\b", r"\bzero[- ]shot\b", r"\bin[- ]context\b"],
    "rag": [r"\brag\b", r"\bretrieval[- ]augmented\b", r"\bdense retrieval\b", r"\bvector database\b"],
    "rlhf": [r"\brlhf\b", r"\bpreference\b", r"\bppo\b", r"\bdpo\b", r"\bipo\b"],
    "alignment": [r"\balignment\b", r"\bsafety\b", r"\bharmless\b", r"\bhelpful\b"],
    "lora_peft": [r"\blora\b", r"\bpeft\b", r"\badapter\b", r"\bprefi(x|ce)[- ]tuning\b"],
    "distillation": [r"\bdistill(ation)?\b", r"\bteacher[- ]student\b"],
    "quantization": [r"\bquantiz(e|ation)\b", r"\b4[- ]bit\b", r"\b8[- ]bit\b", r"\bint8\b"],
    "pruning": [r"\bprun(e|ing)\b"],
    "sparsity_moe": [r"\bspars(e|ity)\b", r"\bmoe\b", r"\bmixture of experts\b"],
    "contrastive": [r"\bcontrastive\b", r"\binfo(nce)?\b", r"\bclip\b"],
    "diffusion": [r"\bdiffusion\b", r"\bscore[- ]based\b"],
    "gan": [r"\bgan\b", r"\bgenerative adversarial\b"],
    "gnn": [r"\bgnn\b", r"\bgraph neural\b", r"\bgcn\b", r"\bgat\b"],
    "privacy_dp": [r"\bdifferential privacy\b", r"\bdp[- ]sgd\b", r"\bprivacy\b"],
    "robustness": [r"\brobust(ness)?\b", r"\badversarial\b", r"\bood\b", r"\bdistribution shift\b"],
    "federated": [r"\bfederated\b", r"\bfedavg\b", r"\bfedprox\b", r"\bfl\b"],
}

TASK_PATTERNS = {
    "qa": [r"\bquestion answering\b", r"\bqa\b", r"\bread(ing)? comprehension\b"],
    "summarization": [r"\bsummarization\b", r"\brouge\b"],
    "mt": [r"\bmachine translation\b", r"\bwmt\b", r"\bbleu\b"],
    "nli": [r"\bnli\b", r"\bentailment\b", r"\bmnli\b"],
    "ner": [r"\bner\b", r"\bnamed entity\b"],
    "sentiment": [r"\bsentiment\b", r"\bsst[- ]?2\b"],
    "retrieval": [r"\bretrieval\b", r"\binformation retrieval\b", r"\bdense retrieval\b"],
    "dialogue": [r"\bdialog(ue)?\b", r"\bchatbot\b"],
    "generation": [r"\btext generation\b", r"\bgenerate\b"],
    "reasoning": [r"\breasoning\b", r"\bmath\b", r"\blogic\b", r"\bchain[- ]of[- ]thought\b"],
    "code": [r"\bcode\b", r"\bhuman[- ]?eval\b", r"\bprogram synthesis\b"],
    "multimodal": [r"\bmultimodal\b", r"\bvision[- ]language\b", r"\bvqa\b"],
}

DATASET_PATTERNS = {
    "glue": [r"\bglue\b"],
    "superglue": [r"\bsuperglue\b"],
    "squad": [r"\bsquad\b"],
    "mnli": [r"\bmnli\b"],
    "sst2": [r"\bsst[- ]?2\b"],
    "wmt": [r"\bwmt\b"],
    "cnn_dailymail": [r"\bcnn[/ -]?dailymail\b"],
    "xsum": [r"\bxsum\b"],
    "mmlu": [r"\bmmlu\b"],
    "gsm8k": [r"\bgsm8k\b"],
    "bigbench": [r"\bbig[- ]?bench\b"],
    "helm": [r"\bhelm\b"],
    "human_eval": [r"\bhuman[- ]?eval\b"],
    "imagenet": [r"\bimagenet\b"],
    "cifar10": [r"\bcifar[- ]?10\b"],
}

METRIC_PATTERNS = {
    "accuracy": [r"\baccuracy\b", r"\bacc\b"],
    "f1": [r"\bf1\b", r"\bf[- ]?score\b"],
    "bleu": [r"\bbleu\b"],
    "rouge": [r"\brouge\b"],
    "em": [r"\bexact match\b", r"\bem\b"],
    "perplexity": [r"\bperplexity\b", r"\bppl\b"],
}

LIMITATION_CUES = [
    r"\blimitation(s)?\b", r"\bwe (do not|don't|cannot|can't)\b", r"\bhowever\b",
    r"\bshortcoming(s)?\b", r"\bweakness(es)?\b", r"\bnot addressed\b", r"\bwe leave\b",
    r"\bconstraints?\b", r"\bdoes not\b", r"\bdoesn't\b"
]
FUTURE_CUES = [
    r"\bfuture work\b", r"\bwe plan to\b", r"\bwe will\b", r"\bin future\b", r"\bnext step\b",
    r"\bwe intend\b", r"\bwill explore\b", r"\bfurther research\b"
]
CONTRIB_CUES = [
    r"\bwe propose\b", r"\bwe present\b", r"\bwe introduce\b", r"\bour contributions?\b",
    r"\bwe develop\b", r"\bwe show\b", r"\bwe demonstrate\b"
]


def match_lexicon(text: str, patterns: Dict[str, List[str]], cap: int = 12) -> List[str]:
    t = normalize_text(text)
    out = []
    for tag, pats in patterns.items():
        for pat in pats:
            if re.search(pat, t):
                out.append(tag)
                break
    return out[:cap]

def extract_by_cues(text: str, cues: List[str], max_sentences: int = 10) -> List[str]:
    out = []
    for s in split_sentences(text):
        ls = s.lower()
        if any(re.search(c, ls) for c in cues):
            out.append(s)
            if len(out) >= max_sentences:
                break
    return dedup_keep_order(out)

def extract_metrics(text: str, max_items: int = 10) -> List[Dict[str, Any]]:
    t = normalize_text(text)
    num = r"([0-9]+(?:\.[0-9]+)?)"
    items = []
    for mname, pats in METRIC_PATTERNS.items():
        for pat in pats:
            rgx1 = re.compile(pat + r"[^0-9]{0,20}" + num)
            for m in rgx1.finditer(t):
                items.append({"name": mname, "value": float(m.group(1))})
                if len(items) >= max_items:
                    break
            if len(items) >= max_items:
                break
    # dedup: keep max value per metric
    best = {}
    for it in items:
        if it["name"] not in best or it["value"] > best[it["name"]]:
            best[it["name"]] = it["value"]
    return [{"name": k, "value": best[k]} for k in sorted(best.keys())][:max_items]


# ============================================================
# TF-IDF + clustering
# ============================================================

def build_tfidf(docs: List[str], min_df: int = 2, max_features: int = 60000) -> np.ndarray:
    tokenized = [tokenize(d) for d in docs]
    df = Counter()
    for toks in tokenized:
        df.update(set(toks))

    vocab = [t for t, c in df.items() if c >= min_df]
    vocab.sort()
    if len(vocab) > max_features:
        vocab = vocab[:max_features]
    tok2idx = {t:i for i,t in enumerate(vocab)}

    N, V = len(docs), len(vocab)
    X = np.zeros((N, V), dtype=np.float32)
    for i, toks in enumerate(tokenized):
        c = Counter([t for t in toks if t in tok2idx])
        if not c:
            continue
        mx = max(c.values())
        for t, v in c.items():
            X[i, tok2idx[t]] = v / mx

    idf = np.zeros((V,), dtype=np.float32)
    for t, j in tok2idx.items():
        idf[j] = math.log((1 + N) / (1 + df[t])) + 1.0

    X *= idf[None, :]
    X /= (np.linalg.norm(X, axis=1, keepdims=True) + 1e-9)
    return X

def choose_k(n: int) -> int:
    if n <= 80:
        return max(4, int(math.sqrt(n)))
    if n <= 800:
        return min(60, max(10, int(math.sqrt(n)*1.2)))
    return min(140, max(25, int(math.sqrt(n)*1.5)))

def kmeans_fallback(X: np.ndarray, k: int, iters: int = 30, seed: int = 13) -> np.ndarray:
    rng = np.random.default_rng(seed)
    n = X.shape[0]
    k = min(max(2, k), n)
    centers = X[rng.choice(n, size=k, replace=False)]
    for _ in range(iters):
        sims = X @ centers.T
        a = np.argmax(sims, axis=1)
        new = np.zeros_like(centers)
        for j in range(k):
            idx = np.where(a == j)[0]
            if len(idx) == 0:
                new[j] = X[rng.integers(0, n)]
            else:
                c = X[idx].mean(axis=0)
                c = c / (np.linalg.norm(c) + 1e-9)
                new[j] = c
        centers = new
    return a.astype(np.int32)

def cluster_topics(X: np.ndarray, k_topics: int) -> np.ndarray:
    try:
        from sklearn.cluster import KMeans
        km = KMeans(n_clusters=k_topics, random_state=13, n_init=10, max_iter=300)
        return km.fit_predict(X).astype(np.int32)
    except Exception:
        return kmeans_fallback(X, k_topics)


# ============================================================
# NEW: Load functions for new JSON formats
# ============================================================

def load_all_sample_json(path: str) -> List[Dict[str, Any]]:
    """
    Load papers from all_sample.json format.
    Structure: [{"query": ..., "data": [paper1, paper2, ...]}]
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    papers = []
    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict) and "data" in item:
                for paper in item.get("data", []):
                    papers.append(paper)
    return papers

def load_imrad_corpus_json(path: str) -> Dict[str, Dict[str, Any]]:
    """
    Load fulltext sections from imrad_corpus.json.
    Structure: [{"paperId": ..., "title": ..., "sections": {...}}]
    Returns: dict mapping paperId -> sections dict
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    corpus = {}
    if isinstance(data, list):
        for paper in data:
            paper_id = paper.get("paperId")
            sections = paper.get("sections", {})
            if paper_id and sections:
                corpus[paper_id] = sections
    return corpus

def normalize_seed_record_new(rec: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normalize paper record from new JSON format (all_sample.json).
    """
    pid = rec.get("paperId") or rec.get("paper_id") or rec.get("id")
    if not pid:
        raise ValueError("Missing paperId")
    
    year = rec.get("year")
    if year is None:
        year = 2020  # default fallback
    
    venue = rec.get("venue", "")
    title = rec.get("title", "")
    abstract = rec.get("abstract", "")
    
    # Extract reference paper IDs (if available in future)
    # For now, we don't have references in the new format
    refs = []
    
    return {
        "paper_id": str(pid),
        "conference": str(venue),
        "year": int(year),
        "title": str(title),
        "abstract": str(abstract),
        "contributions": [],  # Will be extracted from fulltext
        "references": refs,
        "citation_count": rec.get("citationCount", 0),
        "reference_count": rec.get("referenceCount", 0),
    }

def load_fulltext_sections_new(corpus: Dict[str, Dict[str, Any]], paper_id: str) -> Dict[str, str]:
    """
    Load fulltext sections from the imrad_corpus dictionary.
    Returns sections as dict with lowercase keys.
    """
    if paper_id not in corpus:
        return {}
    
    sections_raw = corpus[paper_id]
    sections = {}
    
    for key, value in sections_raw.items():
        key_lower = key.lower()
        if isinstance(value, list):
            # Join list of paragraphs
            sections[key_lower] = "\n\n".join(str(v) for v in value if v)
        elif isinstance(value, str):
            sections[key_lower] = value
    
    return sections


# ============================================================
# Dataset build (updated for new format)
# ============================================================

def emit_events_for_paper(p: Dict[str, Any]) -> List[Dict[str, Any]]:
    unit_id = p["unit_id"]
    ts = str(p["year"])
    payload = {
        "paper_id": p["paper_id"],
        "conference": p.get("conference",""),
        "year": int(p["year"]),
        "title": p.get("title",""),
        "topic_id": p.get("topic_id",""),
        "methods": p.get("method_tags", []),
        "tasks": p.get("task_tags", []),
        "datasets": p.get("dataset_tags", []),
        "metrics": p.get("metrics", []),
    }

    def mk(event_type: str, subtype: str, text: str):
        return {
            "event_id": f"{unit_id}::{p['paper_id']}::{event_type}::{subtype}",
            "track": "ml",
            "unit_id": unit_id,
            "event_type": event_type,
            "subtype": subtype,
            "timestamp": ts,
            "methodology": p.get("method_tags", []),
            "source_type": "paper",
            "text": text,
            "payload": payload
        }

    events = []
    for c in p.get("contributions", [])[:12]:
        events.append(mk("contribution", "model_contribution", c))

    for l in p.get("limitations", [])[:10]:
        events.append(mk("gap", "evaluation_gap", l))

    for f in p.get("future_work", [])[:10]:
        ft = normalize_text(f)
        subtype = "scientific_future_work"
        if any(x in ft for x in ["efficient", "compress", "quantiz", "prun", "speed", "latency"]):
            subtype = "efficiency_future_work"
        elif any(x in ft for x in ["robust", "adversarial", "shift", "ood"]):
            subtype = "robustness_future_work"
        elif "privacy" in ft or "dp" in ft:
            subtype = "privacy_future_work"
        events.append(mk("future", subtype, f))
    return events

def mmr_select(candidates: List[Tuple[int, float]], sim: np.ndarray, k: int, lam: float = 0.75) -> List[int]:
    candidates = sorted(candidates, key=lambda x: x[1], reverse=True)
    chosen = []
    for _ in range(k):
        best_j, best_val = None, -1e9
        for j, score in candidates:
            if j in chosen:
                continue
            div = max(float(sim[j, cj]) for cj in chosen) if chosen else 0.0
            val = lam * score - (1.0 - lam) * div
            if val > best_val:
                best_val = val
                best_j = j
        if best_j is None:
            break
        chosen.append(best_j)
    return chosen

def build_edges_for_topic(topic_papers: List[Dict[str, Any]], topic_X: np.ndarray,
                          top_k: int = 6, max_year_ahead: int = 3,
                          tau_future: float = 0.42, tau_limit: float = 0.42) -> List[Dict[str, Any]]:
    n = len(topic_papers)
    if n <= 1:
        return []
    sim = topic_X @ topic_X.T

    # within-year rank to avoid cycles
    for p in topic_papers:
        p["_rank"] = stable_hash_int(p["paper_id"]) % (10**9)

    edges = []
    w_sim, w_meth, w_task, w_data, w_cite = 0.45, 0.20, 0.10, 0.10, 0.15

    for i, pi in enumerate(topic_papers):
        yi = int(pi["year"])
        candidates = []
        for j, pj in enumerate(topic_papers):
            if i == j:
                continue
            yj = int(pj["year"])
            if yj < yi or yj > yi + max_year_ahead:
                continue
            if yj == yi and not (pi["_rank"] < pj["_rank"]):
                continue

            s = float(sim[i, j])
            meth = jaccard(pi.get("method_tags", []), pj.get("method_tags", []))
            task = jaccard(pi.get("task_tags", []), pj.get("task_tags", []))
            data = jaccard(pi.get("dataset_tags", []), pj.get("dataset_tags", []))
            cite = 1.0 if (pi["paper_id"] in set(pj.get("references", []))) else 0.0
            score = w_sim*s + w_meth*meth + w_task*task + w_data*data + w_cite*cite
            candidates.append((j, score))

        chosen = mmr_select(candidates, sim, k=top_k, lam=0.75)
        for j in chosen:
            pj = topic_papers[j]
            edges.append({
                "edge_id": f"{pi['paper_id']}->({pj['paper_id']})::temporal_related",
                "track": "ml",
                "unit_id": pi["unit_id"],
                "src_paper_id": pi["paper_id"],
                "tgt_paper_id": pj["paper_id"],
                "src_year": yi,
                "tgt_year": int(pj["year"]),
                "edge_type": "temporal_related",
                "score": float(sim[i, j]),
                "evidence": {
                    "citation_supported": bool(pi["paper_id"] in set(pj.get("references", []))),
                    "method_overlap": list(set(pi.get("method_tags", [])) & set(pj.get("method_tags", []))),
                    "task_overlap": list(set(pi.get("task_tags", [])) & set(pj.get("task_tags", []))),
                    "dataset_overlap": list(set(pi.get("dataset_tags", [])) & set(pj.get("dataset_tags", []))),
                }
            })

        # future_realized / limitation_addressed (fast proxy: paper sim)
        if pi.get("future_work"):
            for j, pj in enumerate(topic_papers):
                if i == j:
                    continue
                yj = int(pj["year"])
                if yj < yi or yj > yi + max_year_ahead:
                    continue
                if yj == yi and not (pi["_rank"] < pj["_rank"]):
                    continue
                if float(sim[i, j]) >= tau_future:
                    edges.append({
                        "edge_id": f"{pi['paper_id']}->({pj['paper_id']})::future_realized",
                        "track": "ml", "unit_id": pi["unit_id"],
                        "src_paper_id": pi["paper_id"], "tgt_paper_id": pj["paper_id"],
                        "src_year": yi, "tgt_year": yj,
                        "edge_type": "future_realized",
                        "score": float(sim[i, j]),
                        "evidence": {"future_statements": pi["future_work"][:3]}
                    })
        if pi.get("limitations"):
            for j, pj in enumerate(topic_papers):
                if i == j:
                    continue
                yj = int(pj["year"])
                if yj < yi or yj > yi + max_year_ahead:
                    continue
                if yj == yi and not (pi["_rank"] < pj["_rank"]):
                    continue
                if float(sim[i, j]) >= tau_limit:
                    edges.append({
                        "edge_id": f"{pi['paper_id']}->({pj['paper_id']})::limitation_addressed",
                        "track": "ml", "unit_id": pi["unit_id"],
                        "src_paper_id": pi["paper_id"], "tgt_paper_id": pj["paper_id"],
                        "src_year": yi, "tgt_year": yj,
                        "edge_type": "limitation_addressed",
                        "score": float(sim[i, j]),
                        "evidence": {"limitations": pi["limitations"][:3]}
                    })

        # explicit citation_support
        for j, pj in enumerate(topic_papers):
            if i == j:
                continue
            if pi["paper_id"] in set(pj.get("references", [])):
                yj = int(pj["year"])
                if yj < yi or (yj == yi and not (pi["_rank"] < pj["_rank"])):
                    continue
                edges.append({
                    "edge_id": f"{pi['paper_id']}->({pj['paper_id']})::citation_support",
                    "track": "ml", "unit_id": pi["unit_id"],
                    "src_paper_id": pi["paper_id"], "tgt_paper_id": pj["paper_id"],
                    "src_year": yi, "tgt_year": yj,
                    "edge_type": "citation_support",
                    "score": 1.0,
                    "evidence": {"citation": True}
                })

    # dedup edges by id keep max score
    best = {}
    for e in edges:
        eid = e["edge_id"]
        if eid not in best or e.get("score",0.0) > best[eid].get("score",0.0):
            best[eid] = e
    return list(best.values())

def time_split(papers: List[Dict[str, Any]], val_years: int = 2, test_years: int = 2) -> Tuple[Set[str], Set[str], Set[str]]:
    years = sorted({int(p["year"]) for p in papers})
    if len(years) <= (val_years + test_years + 1):
        ids = [p["paper_id"] for p in papers]
        rng = np.random.default_rng(13)
        rng.shuffle(ids)
        n = len(ids)
        tr = set(ids[:int(0.7*n)])
        va = set(ids[int(0.7*n):int(0.85*n)])
        te = set(ids[int(0.85*n):])
        return tr, va, te
    testY = set(years[-test_years:])
    valY = set(years[-(test_years+val_years):-test_years])
    tr, va, te = set(), set(), set()
    for p in papers:
        y = int(p["year"])
        if y in testY: te.add(p["paper_id"])
        elif y in valY: va.add(p["paper_id"])
        else: tr.add(p["paper_id"])
    return tr, va, te


def _find_all_occurrences(haystack: str, needle: str) -> List[Tuple[int,int]]:
    """Case-insensitive substring matches; returns list of (start,end)."""
    H = haystack.lower()
    N = needle.lower().strip()
    if not N:
        return []
    out = []
    start = 0
    while True:
        k = H.find(N, start)
        if k == -1:
            break
        out.append((k, k + len(N)))
        start = k + 1
    return out

def _nearest_block_boundary(text: str, i: int, direction: int) -> int:
    """
    Expand to paragraph boundary using blank-line separation.
    direction: -1 for left, +1 for right
    """
    if direction < 0:
        j = i
        while j > 0:
            if text[j-1] == "\n" and j-2 >= 0 and text[j-2] == "\n":
                break
            j -= 1
        return j
    else:
        j = i
        n = len(text)
        while j < n:
            if j+1 < n and text[j] == "\n" and text[j+1] == "\n":
                j += 2
                break
            j += 1
        return min(j, n)

def reconstruct_segments_for_paper(
    paper: Dict[str, Any],
    gap_chars: int = 350,
    expand_to_paragraph: bool = True,
    section_priority: Optional[Dict[str, List[str]]] = None
) -> Tuple[List[Dict[str, Any]], Dict[str, str]]:
    """
    Returns:
      segments: list of segment dicts
      sent2seg: mapping from (label_type, sentence_hash) -> segment_id
    Requires paper["sections"] to exist for full grounding. If absent, segments will be empty.
    """
    sections = paper.get("sections", {})
    if not isinstance(sections, dict) or not sections:
        return [], {}

    # Which sections are allowed per label_type (you can tune these)
    if section_priority is None:
        section_priority = {
            "contribution": ["introduction", "abstract", "conclusion"],
            "limitation": ["discussion", "conclusion"],
            "future_work": ["discussion", "conclusion"],
        }

    def sentence_key(label_type: str, sent: str) -> str:
        return f"{label_type}::{stable_hash_int(sent)}"

    # Gather sentence spans per (label_type, section)
    spans = defaultdict(list)  # (label_type, section) -> list of dicts {start,end,sentence}
    ungrounded = []            # keep track if needed
    sent2seg = {}              # label_type::hash -> segment_id (filled later)

    for label_type, field in [("contribution","contributions"),
                              ("limitation","limitations"),
                              ("future_work","future_work")]:
        sents = paper.get(field, [])
        if not isinstance(sents, list) or not sents:
            continue
        sec_list = section_priority.get(label_type, [])

        for sent in sents:
            sent = (sent or "").strip()
            if not sent:
                continue
            found = False
            for sec in sec_list:
                T = sections.get(sec, "")
                if not isinstance(T, str) or not T.strip():
                    continue
                occ = _find_all_occurrences(T, sent)
                if occ:
                    # choose earliest occurrence (deterministic)
                    st, en = occ[0]
                    spans[(label_type, sec)].append({"start": st, "end": en, "sentence": sent})
                    found = True
                    break
            if not found:
                ungrounded.append({"label_type": label_type, "sentence": sent})

    # Merge spans into segments
    segments = []
    seg_counter = 0

    for (label_type, sec), items in spans.items():
        T = sections.get(sec, "")
        items = sorted(items, key=lambda x: (x["start"], x["end"]))

        merged = []
        cur = None
        for it in items:
            if cur is None:
                cur = {"start": it["start"], "end": it["end"], "sentences": [it["sentence"]]}
                continue
            if it["start"] - cur["end"] <= gap_chars:
                cur["end"] = max(cur["end"], it["end"])
                cur["sentences"].append(it["sentence"])
            else:
                merged.append(cur)
                cur = {"start": it["start"], "end": it["end"], "sentences": [it["sentence"]]}
        if cur is not None:
            merged.append(cur)

        # optionally expand merged spans to paragraph boundaries
        for m in merged:
            st, en = m["start"], m["end"]
            if expand_to_paragraph:
                st = _nearest_block_boundary(T, st, -1)
                en = _nearest_block_boundary(T, en, +1)
            seg_text = T[st:en].strip()

            seg_id = f"ml::SEG::{paper['paper_id']}::{label_type}::{sec}::{seg_counter:03d}"
            seg_counter += 1

            seg_obj = {
                "segment_id": seg_id,
                "track": "ml",
                "paper_id": paper["paper_id"],
                "unit_id": paper.get("unit_id",""),
                "topic_id": paper.get("topic_id",""),
                "year": int(paper.get("year", 0)),
                "conference": paper.get("conference",""),
                "label_type": label_type,
                "section": sec,
                "span_start": int(st),
                "span_end": int(en),
                "text": seg_text,
                "sentences": dedup_keep_order(m["sentences"]),
                "derivation": {
                    "method": "substring_grounding+gap_merge+paragraph_expand" if expand_to_paragraph else "substring_grounding+gap_merge",
                    "gap_chars": gap_chars,
                    "expand_to_paragraph": expand_to_paragraph,
                    "section_priority": section_priority.get(label_type, [])
                }
            }
            segments.append(seg_obj)

            # map sentences -> segment_id
            for s in seg_obj["sentences"]:
                sent2seg[sentence_key(label_type, s)] = seg_id

    return segments, sent2seg


def emit_segment_events(segments: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Segment-level events (parallel to sentence events).
    Stored in evobench_ml_segment_events.jsonl
    """
    events = []
    for seg in segments:
        unit_id = seg.get("unit_id","")
        paper_id = seg["paper_id"]
        ts = str(seg.get("year",""))
        # normalize subtype similar to sentence-level
        if seg["label_type"] == "contribution":
            ev_type, subtype = "contribution", "segment_contribution"
        elif seg["label_type"] == "limitation":
            ev_type, subtype = "gap", "segment_limitation"
        else:
            ev_type, subtype = "future", "segment_future_work"

        events.append({
            "event_id": f"{unit_id}::{paper_id}::SEG::{seg['segment_id']}",
            "track": "ml",
            "unit_id": unit_id,
            "event_type": ev_type,
            "subtype": subtype,
            "timestamp": ts,
            "source_type": "paper_segment",
            "text": seg["text"],
            "payload": {
                "paper_id": paper_id,
                "segment_id": seg["segment_id"],
                "label_type": seg["label_type"],
                "section": seg["section"],
                "span_start": seg["span_start"],
                "span_end": seg["span_end"],
                "topic_id": seg.get("topic_id",""),
                "conference": seg.get("conference",""),
                "year": int(seg.get("year",0)),
            }
        })
    return events


def cmd_build(args):
    ensure_dir(args.out_dir)
    ensure_dir(os.path.join(args.out_dir, "splits"))

    # Load papers from new JSON format
    print(f"Loading papers from {args.seed}...")
    raw_papers = load_all_sample_json(args.seed)
    print(f"Loaded {len(raw_papers)} papers from seed file")

    # Load fulltext corpus if provided
    fulltext_corpus = {}
    if args.fulltext_file:
        print(f"Loading fulltext from {args.fulltext_file}...")
        fulltext_corpus = load_imrad_corpus_json(args.fulltext_file)
        print(f"Loaded fulltext for {len(fulltext_corpus)} papers")

    # Process ALL papers - no filtering
    papers = []
    skipped = 0
    for rec in raw_papers:
        try:
            p = normalize_seed_record_new(rec)
            papers.append(p)
        except Exception as e:
            print(f"Warning: Skipping paper due to error: {e}")
            skipped += 1
            continue
    
    if skipped > 0:
        print(f"Note: Skipped {skipped} papers due to errors")
    
    if len(papers) < 20:
        raise RuntimeError(f"Not enough valid papers. Got {len(papers)}, need at least 20.")
    
    print(f"Processing ALL {len(papers)} papers...")

    # attach fulltext and extract features
    for i, p in enumerate(papers):
        if (i + 1) % 100 == 0:
            print(f"  Processing paper {i+1}/{len(papers)}...")
        
        secs = load_fulltext_sections_new(fulltext_corpus, p["paper_id"])
        intro = secs.get("introduction","")
        methods = secs.get("methods","")
        results = secs.get("results","")
        disc = secs.get("discussion","")
        concl = secs.get("conclusion","")

        # contributions fallback
        if not p["contributions"]:
            p["contributions"] = extract_by_cues((p["abstract"]+"\n"+intro), CONTRIB_CUES, 8)

        # tags
        meth_src = methods if methods.strip() else (p["title"]+" "+p["abstract"])
        tds_src = (p["abstract"]+"\n"+intro+"\n"+methods)

        p["method_tags"] = match_lexicon(meth_src, METHOD_PATTERNS, cap=12)
        p["task_tags"] = match_lexicon(tds_src, TASK_PATTERNS, cap=10)
        p["dataset_tags"] = match_lexicon(tds_src, DATASET_PATTERNS, cap=10)
        p["metrics"] = extract_metrics(results, 10) if results.strip() else []

        # limitation/future: discussion+conclusion only
        lf_src = (disc + "\n" + concl).strip()
        p["limitations"] = extract_by_cues(lf_src, LIMITATION_CUES, 10) if lf_src else []
        p["future_work"] = extract_by_cues(lf_src, FUTURE_CUES, 10) if lf_src else []

        if args.store_sections:
            p["sections"] = secs

    print("Building TF-IDF representations...")
    # topic clustering
    docs = []
    for p in papers:
        c = " ".join(p.get("contributions", [])[:6])
        lf = " ".join(p.get("limitations", [])[:3] + p.get("future_work", [])[:3])
        docs.append(f"{p['title']} {p['abstract']} {c} {lf}")

    X = build_tfidf(docs, min_df=args.min_df, max_features=args.max_features)
    k = args.k_topics if args.k_topics > 0 else choose_k(len(papers))
    
    print(f"Clustering into {k} topics...")
    assign = cluster_topics(X, k)

    topic2idx = defaultdict(list)
    units = []
    for i, p in enumerate(papers):
        tid = int(assign[i])
        unit_id = f"ml::T{tid:03d}"
        p["topic_id"] = f"T{tid:03d}"
        p["unit_id"] = unit_id
        p["_rank"] = stable_hash_int(p["paper_id"]) % (10**9)
        topic2idx[unit_id].append(i)

    for uid, idxs in topic2idx.items():
        units.append({
            "unit_id": uid,
            "track": "ml",
            "name": uid,
            "metadata": {"topic_size": len(idxs), "source": "semantic_scholar+imrad heuristics"}
        })

    print(f"Building edges...")
    # edges
    edges = []
    for uid, idxs in topic2idx.items():
        tp = [papers[i] for i in idxs]
        tX = X[idxs]
        edges.extend(build_edges_for_topic(tp, tX, top_k=args.top_k_edges,
                                          max_year_ahead=args.max_year_ahead,
                                          tau_future=args.tau_future,
                                          tau_limit=args.tau_limit))

    print(f"Generating events...")
    # events
    events = []
    for p in papers:
        events.extend(emit_events_for_paper(p))

    # Segment-level annotations
    print(f"Reconstructing text segments...")
    all_segments = []
    all_seg_events = []
    for p in papers:
        segs, sent2seg = reconstruct_segments_for_paper(
            p,
            gap_chars=args.segment_gap_chars,
            expand_to_paragraph=(args.segment_expand_to_paragraph.lower() == "true")
        )
        p["sentence_to_segment"] = sent2seg
        all_segments.extend(segs)
        all_seg_events.extend(emit_segment_events(segs))

    print(f"Writing output files...")
    write_jsonl(os.path.join(args.out_dir, "evobench_ml_segments.jsonl"), all_segments)
    write_jsonl(os.path.join(args.out_dir, "evobench_ml_segment_events.jsonl"), all_seg_events)

    # write
    write_jsonl(os.path.join(args.out_dir, "raw_ml_papers.jsonl"), papers)
    write_jsonl(os.path.join(args.out_dir, "evobench_ml_units.jsonl"), units)
    write_jsonl(os.path.join(args.out_dir, "evobench_ml_edges.jsonl"), edges)
    write_jsonl(os.path.join(args.out_dir, "evobench_ml_events.jsonl"), events)

    # splits
    print(f"Creating train/val/test splits...")
    tr, va, te = time_split(papers, args.val_years, args.test_years)
    splits = os.path.join(args.out_dir, "splits")
    write_jsonl(os.path.join(splits, "train_papers.jsonl"), [p for p in papers if p["paper_id"] in tr])
    write_jsonl(os.path.join(splits, "val_papers.jsonl"),   [p for p in papers if p["paper_id"] in va])
    write_jsonl(os.path.join(splits, "test_papers.jsonl"),  [p for p in papers if p["paper_id"] in te])
    write_jsonl(os.path.join(splits, "train_events.jsonl"), [e for e in events if e["payload"]["paper_id"] in tr])
    write_jsonl(os.path.join(splits, "val_events.jsonl"),   [e for e in events if e["payload"]["paper_id"] in va])
    write_jsonl(os.path.join(splits, "test_events.jsonl"),  [e for e in events if e["payload"]["paper_id"] in te])

    # only keep intra-split edges for clean train/val/test graphs
    write_jsonl(os.path.join(splits, "train_edges.jsonl"), [e for e in edges if e["src_paper_id"] in tr and e["tgt_paper_id"] in tr])
    write_jsonl(os.path.join(splits, "val_edges.jsonl"),   [e for e in edges if e["src_paper_id"] in va and e["tgt_paper_id"] in va])
    write_jsonl(os.path.join(splits, "test_edges.jsonl"),  [e for e in edges if e["src_paper_id"] in te and e["tgt_paper_id"] in te])

    print("\n" + "="*60)
    print("BUILD COMPLETE")
    print("="*60)
    print(f"Output directory: {args.out_dir}")
    print(f"Papers: {len(papers)}")
    print(f"Topics: {len(units)}")
    print(f"Events: {len(events)}")
    print(f"Edges: {len(edges)}")
    print(f"Segments: {len(all_segments)}")
    print(f"Segment events: {len(all_seg_events)}")
    print(f"\nSplits:")
    print(f"  Train: {len(tr)} papers")
    print(f"  Val:   {len(va)} papers")
    print(f"  Test:  {len(te)} papers")
    print("="*60)


# ============================================================
# Automated validation++ (no changes needed)
# ============================================================

def validate_temporal(edges: List[Dict[str, Any]], papers_by_id: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    violations = 0
    same_year_rank_viol = 0
    for e in edges:
        s, t = e["src_paper_id"], e["tgt_paper_id"]
        if s not in papers_by_id or t not in papers_by_id:
            continue
        ys, yt = int(papers_by_id[s]["year"]), int(papers_by_id[t]["year"])
        if yt < ys:
            violations += 1
        if yt == ys:
            rs, rt = papers_by_id[s].get("_rank"), papers_by_id[t].get("_rank")
            if rs is not None and rt is not None and not (rs < rt):
                same_year_rank_viol += 1
    return {"temporal_violations": violations, "same_year_rank_violations": same_year_rank_viol}

def edge_triangulation(edges: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Edge support by multiple signals: citation + overlaps + similarity score
    """
    supported2 = 0
    supported0 = 0
    total = 0
    for e in edges:
        if e.get("edge_type") not in ["temporal_related", "future_realized", "limitation_addressed"]:
            continue
        total += 1
        ev = e.get("evidence", {})
        cite = 1 if ev.get("citation_supported") or ev.get("citation") else 0
        overlaps = 1 if (len(ev.get("method_overlap", [])) + len(ev.get("task_overlap", [])) + len(ev.get("dataset_overlap", [])) > 0) else 0
        sim = 1 if float(e.get("score", 0.0)) >= 0.35 else 0  # coarse
        s = cite + overlaps + sim
        if s >= 2:
            supported2 += 1
        if s == 0:
            supported0 += 1
    return {
        "triangulated_edges_total": total,
        "supported_by_2plus": supported2,
        "supported_by_2plus_rate": supported2 / max(1, total),
        "supported_by_0": supported0,
        "supported_by_0_rate": supported0 / max(1, total),
    }

def future_realization(papers: List[Dict[str, Any]], edges: List[Dict[str, Any]]) -> Dict[str, Any]:
    has_future = set(p["paper_id"] for p in papers if p.get("future_work"))
    has_lim = set(p["paper_id"] for p in papers if p.get("limitations"))
    fut_real = set(e["src_paper_id"] for e in edges if e.get("edge_type") == "future_realized")
    lim_addr = set(e["src_paper_id"] for e in edges if e.get("edge_type") == "limitation_addressed")
    fr = len(fut_real & has_future) / max(1, len(has_future))
    lr = len(lim_addr & has_lim) / max(1, len(has_lim))
    return {
        "papers_with_future": len(has_future),
        "papers_with_future_realized": len(fut_real & has_future),
        "future_realization_rate": fr,
        "future_realization_ci95": list(wilson_ci(len(fut_real & has_future), max(1, len(has_future)))),
        "papers_with_limitations": len(has_lim),
        "papers_with_limitation_addressed": len(lim_addr & has_lim),
        "limitation_address_rate": lr,
        "limitation_address_ci95": list(wilson_ci(len(lim_addr & has_lim), max(1, len(has_lim)))),
    }

def shuffled_time_control(papers: List[Dict[str, Any]], seed: int = 13) -> Dict[str, Any]:
    """
    Shuffle years *within topic* to show temporal structure is meaningful.
    Returns distribution shift summary: same-year proportion changes, etc.
    """
    rng = np.random.default_rng(seed)
    # group by unit
    unit2 = defaultdict(list)
    for p in papers:
        unit2[p["unit_id"]].append(p)
    swapped = 0
    total = 0
    for _, ps in unit2.items():
        years = [p["year"] for p in ps]
        perm = years.copy()
        rng.shuffle(perm)
        for p, ny in zip(ps, perm):
            total += 1
            if int(p["year"]) != int(ny):
                swapped += 1
    return {"shuffled_time_fraction_changed": swapped / max(1, total)}

def cmd_validatepp(args):
    data_dir = args.data_dir
    papers = list(read_jsonl(os.path.join(data_dir, "raw_ml_papers.jsonl")))
    edges = list(read_jsonl(os.path.join(data_dir, "evobench_ml_edges.jsonl")))
    events = list(read_jsonl(os.path.join(data_dir, "evobench_ml_events.jsonl")))

    papers_by_id = {p["paper_id"]: p for p in papers}

    report = {
        "counts": {
            "papers": len(papers),
            "events": len(events),
            "edges": len(edges),
            "topics": len({p["unit_id"] for p in papers})
        },
        "temporal": validate_temporal(edges, papers_by_id),
        "triangulation": edge_triangulation(edges),
        "future_realization": future_realization(papers, edges),
        "shuffled_time_control": shuffled_time_control(papers),
        "edge_types": dict(Counter(e.get("edge_type","unknown") for e in edges)),
        "event_types": dict(Counter(e.get("event_type","unknown") for e in events)),
        "subtype_top10": Counter(e.get("subtype","unknown") for e in events).most_common(10),
    }

    out_path = os.path.join(data_dir, "validation_report_pp.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)
    print(f"VALIDATE++ COMPLETE: {out_path}")
    print(json.dumps(report, indent=2))


# ============================================================
# Human review pack (CSV + XLSX) - no changes needed
# ============================================================

def sample_items_mixed(items: List[Any], scores: Optional[List[float]], n: int, seed: int = 13) -> List[int]:
    """
    Mixed sampling: 70% lowest-score + 30% random
    """
    rng = np.random.default_rng(seed)
    N = len(items)
    if N == 0:
        return []
    idx = np.arange(N)
    if scores is None:
        rng.shuffle(idx)
        return idx[:min(n, N)].tolist()

    scores = np.array(scores, dtype=np.float32)
    order = np.argsort(scores)  # low -> high
    n_low = int(0.70 * n)
    n_rand = n - n_low
    low = order[:min(n_low, N)].tolist()
    remaining = [i for i in idx.tolist() if i not in set(low)]
    rng.shuffle(remaining)
    return (low + remaining[:min(n_rand, len(remaining))])[:min(n, N)]

def autosize_ws(ws):
    for col in ws.columns:
        max_len = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(70, max_len + 2)

def cmd_review_pack(args):
    data_dir = args.data_dir
    out_dir = args.out_dir
    ensure_dir(out_dir)

    papers = list(read_jsonl(os.path.join(data_dir, "raw_ml_papers.jsonl")))
    edges = list(read_jsonl(os.path.join(data_dir, "evobench_ml_edges.jsonl")))

    # Build extraction items
    extraction_rows = []
    extraction_scores = []
    for p in papers:
        pid = p["paper_id"]
        meta = {"paper_id": pid, "year": p["year"], "conference": p.get("conference",""), "topic": p.get("unit_id","")}
        # sample up to 2 each type
        for t, lst in [("contribution", p.get("contributions", [])),
                       ("limitation", p.get("limitations", [])),
                       ("future_work", p.get("future_work", []))]:
            for s in (lst[:2] if isinstance(lst, list) else []):
                extraction_rows.append({
                    **meta,
                    "item_type": t,
                    "extracted_text": s,
                    "context_hint": "discussion/conclusion"
                })
                # heuristic "risk score" for sampling: fewer tags + shorter text => lower confidence
                conf = 0.0
                conf += 0.2 * len(p.get("method_tags", []))
                conf += 0.1 * len(p.get("task_tags", []))
                conf += 0.1 * len(p.get("dataset_tags", []))
                conf += 0.001 * len(s)
                extraction_scores.append(float(conf))

    # Filter edges for review
    edge_rows = []
    edge_scores = []
    for e in edges:
        if e.get("edge_type") not in ["temporal_related", "future_realized", "limitation_addressed"]:
            continue
        edge_rows.append({
            "edge_id": e.get("edge_id",""),
            "edge_type": e.get("edge_type",""),
            "score": float(e.get("score", 0.0)),
            "unit_id": e.get("unit_id",""),
            "src_paper_id": e.get("src_paper_id",""),
            "tgt_paper_id": e.get("tgt_paper_id",""),
            "src_year": e.get("src_year",""),
            "tgt_year": e.get("tgt_year",""),
            "citation_supported": bool(e.get("evidence", {}).get("citation_supported", False)),
            "method_overlap": ", ".join(e.get("evidence", {}).get("method_overlap", [])),
            "task_overlap": ", ".join(e.get("evidence", {}).get("task_overlap", [])),
            "dataset_overlap": ", ".join(e.get("evidence", {}).get("dataset_overlap", [])),
            "src_statement_hint": "see evidence.future_statements/limitations" if e.get("edge_type") != "temporal_related" else "",
        })
        # risk score: low similarity + no overlap + no citation => low confidence
        ov = 0
        ev = e.get("evidence", {})
        ov += len(ev.get("method_overlap", []))
        ov += len(ev.get("task_overlap", []))
        ov += len(ev.get("dataset_overlap", []))
        cite = 1 if ev.get("citation_supported") else 0
        risk_conf = float(e.get("score", 0.0)) + 0.15 * ov + 0.25 * cite
        edge_scores.append(risk_conf)

    # sample indices
    if args.strategy == "mixed":
        ex_idx = sample_items_mixed(extraction_rows, extraction_scores, args.n_extractions, seed=13)
        ed_idx = sample_items_mixed(edge_rows, edge_scores, args.n_edges, seed=13)
    else:
        rng = np.random.default_rng(13)
        ex_idx = rng.choice(len(extraction_rows), size=min(args.n_extractions, len(extraction_rows)), replace=False).tolist()
        ed_idx = rng.choice(len(edge_rows), size=min(args.n_edges, len(edge_rows)), replace=False).tolist()

    ex_sample = [extraction_rows[i] for i in ex_idx]
    ed_sample = [edge_rows[i] for i in ed_idx]

    # Write CSV
    ex_csv = os.path.join(out_dir, "review_extractions.csv")
    ed_csv = os.path.join(out_dir, "review_edges.csv")

    def write_csv(path, rows):
        import csv
        if not rows:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) + [
                "QA_grounded", "QA_correct_type", "QA_specific",
                "QA_temporal", "QA_plausible", "QA_edge_type_correct",
                "QA_notes"
            ])
            w.writeheader()
            for r in rows:
                rr = dict(r)
                rr.update({
                    "QA_grounded": "", "QA_correct_type": "", "QA_specific": "",
                    "QA_temporal": "", "QA_plausible": "", "QA_edge_type_correct": "",
                    "QA_notes": ""
                })
                w.writerow(rr)

    write_csv(ex_csv, ex_sample)
    write_csv(ed_csv, ed_sample)

    # Write XLSX (two sheets)
    xlsx_path = os.path.join(out_dir, "review_pack.xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Extractions"

    def write_sheet(ws, rows):
        if not rows:
            return
        headers = list(rows[0].keys()) + [
            "QA_grounded", "QA_correct_type", "QA_specific",
            "QA_temporal", "QA_plausible", "QA_edge_type_correct",
            "QA_notes"
        ]
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        autosize_ws(ws)

    write_sheet(ws1, ex_sample)
    ws2 = wb.create_sheet("Edges")
    write_sheet(ws2, ed_sample)

    wb.save(xlsx_path)
    print("REVIEW PACK WRITTEN")
    print(f"- {ex_csv}")
    print(f"- {ed_csv}")
    print(f"- {xlsx_path}")


# ============================================================
# Open-source LLM judge (zero-shot, cached) - no changes needed
# ============================================================

def _load_llm(model_name: str, device: str = "auto"):
    if not _TRANSFORMERS_OK:
        raise RuntimeError("transformers/torch not installed. pip install transformers torch")
    tok = AutoTokenizer.from_pretrained(model_name, use_fast=True)
    kwargs = {}
    if device == "auto":
        kwargs["device_map"] = "auto"
    model = AutoModelForCausalLM.from_pretrained(model_name, torch_dtype="auto", **kwargs)
    model.eval()
    return tok, model

def _generate(tok, model, prompt: str, max_new_tokens: int = 96, temperature: float = 0.0):
    inputs = tok(prompt, return_tensors="pt")
    if hasattr(model, "device"):
        inputs = {k: v.to(model.device) for k, v in inputs.items()}
    with torch.no_grad():
        out = model.generate(
            **inputs,
            max_new_tokens=max_new_tokens,
            do_sample=(temperature > 0),
            temperature=temperature if temperature > 0 else None,
            top_p=0.95 if temperature > 0 else None,
            eos_token_id=tok.eos_token_id
        )
    text = tok.decode(out[0], skip_special_tokens=True)
    # return only completion after prompt when possible
    if text.startswith(prompt):
        return text[len(prompt):].strip()
    return text.strip()

def _parse_judge_json(text: str) -> Dict[str, Any]:
    """
    Expect LLM output JSON like:
    {"label":"YES|NO|UNCLEAR","confidence":0.0-1.0,"rationale":"..."}
    """
    # try to find JSON object in text
    m = re.search(r"\{.*\}", text, flags=re.S)
    if not m:
        return {"label": "UNCLEAR", "confidence": 0.0, "rationale": "no_json"}
    try:
        obj = json.loads(m.group(0))
        lab = str(obj.get("label","UNCLEAR")).upper()
        if lab not in ["YES","NO","UNCLEAR"]:
            lab = "UNCLEAR"
        conf = float(obj.get("confidence", 0.0))
        conf = clamp(conf, 0.0, 1.0)
        rat = str(obj.get("rationale","")).strip()
        return {"label": lab, "confidence": conf, "rationale": rat[:300]}
    except Exception:
        return {"label": "UNCLEAR", "confidence": 0.0, "rationale": "json_parse_error"}

def _judge_prompt_extraction(item_type: str, extracted_text: str, context_hint: str = "") -> str:
    return f"""You are validating an information extraction from a research paper.
Task: decide whether the extracted sentence is truly a {item_type} statement.
Return JSON only: {{"label":"YES|NO|UNCLEAR","confidence":0.0-1.0,"rationale":"short"}}

Definitions:
- contribution: what the paper claims it adds/proposes/presents.
- limitation: stated weakness, constraint, missing coverage, or caveat.
- future_work: stated plan or suggested next direction.

Context hint: {context_hint}
Extracted sentence:
\"\"\"{extracted_text}\"\"\"
"""

def _judge_prompt_edge(edge_type: str, src_text: str, tgt_text: str, meta: str = "") -> str:
    return f"""You are validating a temporal relationship between two papers.
Edge type: {edge_type}
Decide if the target text plausibly supports/realizes/addresses the source text given the edge type.
Return JSON only: {{"label":"YES|NO|UNCLEAR","confidence":0.0-1.0,"rationale":"short"}}

Edge semantics:
- temporal_related: same topic and plausible related progress.
- future_realized: target contribution appears to realize the source future work.
- limitation_addressed: target contribution appears to address the source limitation.

Meta: {meta}

Source statement:
\"\"\"{src_text}\"\"\"

Target contribution/summary:
\"\"\"{tgt_text}\"\"\"
"""

def cmd_llm_judge(args):
    data_dir = args.data_dir
    out_file = args.out_file or os.path.join(data_dir, "llm_judgments.jsonl")
    cache_file = args.cache_file or os.path.join(data_dir, "llm_judge_cache.json")

    papers = list(read_jsonl(os.path.join(data_dir, "raw_ml_papers.jsonl")))
    edges = list(read_jsonl(os.path.join(data_dir, "evobench_ml_edges.jsonl")))

    papers_by_id = {p["paper_id"]: p for p in papers}

    cache = {}
    if os.path.exists(cache_file):
        try:
            cache = json.load(open(cache_file, "r", encoding="utf-8"))
        except Exception:
            cache = {}

    tok, model = _load_llm(args.model, args.device)

    judgments = []
    def cached_call(key: str, prompt: str):
        if key in cache:
            return cache[key]
        text = _generate(tok, model, prompt, max_new_tokens=args.max_new_tokens, temperature=args.temperature)
        obj = _parse_judge_json(text)
        cache[key] = obj
        return obj

    # Judge extractions
    if args.judge_extractions.lower() == "true":
        for p in papers:
            pid = p["paper_id"]
            # judge small sample per paper (cap for cost)
            for t, lst in [("contribution", p.get("contributions", [])[:2]),
                           ("limitation", p.get("limitations", [])[:2]),
                           ("future_work", p.get("future_work", [])[:2])]:
                for s in lst:
                    key = f"EX::{pid}::{t}::{stable_hash_int(s)}"
                    prompt = _judge_prompt_extraction(t, s, context_hint="paper-level extraction")
                    res = cached_call(key, prompt)
                    judgments.append({
                        "kind": "extraction",
                        "paper_id": pid,
                        "item_type": t,
                        "text": s,
                        "label": res["label"],
                        "confidence": res["confidence"],
                        "rationale": res["rationale"],
                        "model": args.model
                    })

    # Judge edges
    if args.judge_edges.lower() == "true":
        for e in edges:
            et = e.get("edge_type","")
            if et not in ["temporal_related", "future_realized", "limitation_addressed"]:
                continue
            sp = papers_by_id.get(e["src_paper_id"], {})
            tp = papers_by_id.get(e["tgt_paper_id"], {})
            if not sp or not tp:
                continue

            # source statement: choose from evidence if present
            src_text = ""
            if et == "future_realized":
                src_text = " ".join(e.get("evidence", {}).get("future_statements", [])[:2])
            elif et == "limitation_addressed":
                src_text = " ".join(e.get("evidence", {}).get("limitations", [])[:2])
            else:
                # for temporal_related, use compact signature: title + methods
                src_text = f"{sp.get('title','')} | methods={', '.join(sp.get('method_tags', [])[:5])}"

            # target summary: contribution text if available else title+abstract
            tgt_text = ""
            if tp.get("contributions"):
                tgt_text = " ".join(tp["contributions"][:2])
            else:
                tgt_text = f"{tp.get('title','')} {tp.get('abstract','')[:400]}"

            meta = f"src_year={e.get('src_year')} tgt_year={e.get('tgt_year')} score={e.get('score',0.0):.3f}"
            key = f"ED::{e.get('edge_id')}::{stable_hash_int(src_text+tgt_text)}"
            prompt = _judge_prompt_edge(et, src_text, tgt_text, meta=meta)
            res = cached_call(key, prompt)
            judgments.append({
                "kind": "edge",
                "edge_id": e.get("edge_id",""),
                "edge_type": et,
                "src_paper_id": e["src_paper_id"],
                "tgt_paper_id": e["tgt_paper_id"],
                "label": res["label"],
                "confidence": res["confidence"],
                "rationale": res["rationale"],
                "model": args.model
            })

    # Write judgments
    write_jsonl(out_file, judgments)
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(cache, f)

    print("LLM JUDGING COMPLETE")
    print(f"- judgments: {out_file}")
    print(f"- cache:     {cache_file}")
    print(f"- count:     {len(judgments)}")


# ============================================================
# Stats: ambiguity + agreement + significance - no changes needed
# ============================================================

def cmd_stats(args):
    data_dir = args.data_dir
    judge_file = args.judge_file
    if not os.path.exists(judge_file):
        raise FileNotFoundError(judge_file)

    jud = list(read_jsonl(judge_file))

    # Extraction validity stats
    ex = [j for j in jud if j.get("kind") == "extraction"]
    ed = [j for j in jud if j.get("kind") == "edge"]

    def label_counts(items):
        c = Counter([x.get("label","UNCLEAR") for x in items])
        tot = sum(c.values())
        return {"total": tot, **{k: int(v) for k,v in c.items()},
                "yes_rate": c.get("YES",0)/max(1,tot),
                "no_rate": c.get("NO",0)/max(1,tot),
                "unclear_rate": c.get("UNCLEAR",0)/max(1,tot)}

    report = {
        "extraction_judge": label_counts(ex),
        "edge_judge": label_counts(ed),
    }

    # If you also have a HUMAN review file filled out, compute kappa
    if args.human_csv and os.path.exists(args.human_csv):
        import csv
        human_labels = []
        llm_labels = []
        with open(args.human_csv, "r", encoding="utf-8") as f:
            r = csv.DictReader(f)
            # Expect columns QA_correct_type for extraction or QA_plausible for edges
            for row in r:
                pid = row.get("paper_id","")
                txt = row.get("extracted_text","")
                if pid and txt:
                    # find matching LLM judgment
                    key = (pid, normalize_text(txt))
                    # lookup naive
                    match = None
                    for j in ex:
                        if j.get("paper_id")==pid and normalize_text(j.get("text",""))==key[1]:
                            match = j
                            break
                    if match:
                        h = (row.get("QA_correct_type","") or "").strip().upper()
                        # map human to YES/NO/UNCLEAR
                        if h in ["YES","Y","1","TRUE"]:
                            human_labels.append("YES")
                        elif h in ["NO","N","0","FALSE"]:
                            human_labels.append("NO")
                        else:
                            human_labels.append("UNCLEAR")
                        llm_labels.append(match.get("label","UNCLEAR"))
        if human_labels and llm_labels:
            report["agreement"] = {
                "cohens_kappa": cohens_kappa(human_labels, llm_labels),
                "n_compared": len(human_labels)
            }

    out_path = os.path.join(data_dir, "llm_stats_report.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)
    print(f"STATS COMPLETE: {out_path}")
    print(json.dumps(report, indent=2))


# ============================================================
# CLI (updated for new input format)
# ============================================================

def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build")
    b.add_argument("--seed", required=True, help="Path to all_sample.json")
    b.add_argument("--fulltext_file", default="", help="Path to imrad_corpus.json (optional)")
    b.add_argument("--out_dir", required=True)
    b.add_argument("--min_df", type=int, default=2)
    b.add_argument("--max_features", type=int, default=60000)
    b.add_argument("--k_topics", type=int, default=0)
    b.add_argument("--top_k_edges", type=int, default=6)
    b.add_argument("--max_year_ahead", type=int, default=3)
    b.add_argument("--tau_future", type=float, default=0.42)
    b.add_argument("--tau_limit", type=float, default=0.42)
    b.add_argument("--val_years", type=int, default=2)
    b.add_argument("--test_years", type=int, default=2)
    b.add_argument("--store_sections", action="store_true")
    b.add_argument("--segment_gap_chars", type=int, default=350)
    b.add_argument("--segment_expand_to_paragraph", default="true", choices=["true","false"])

    vpp = sub.add_parser("validatepp")
    vpp.add_argument("--data_dir", required=True)

    rp = sub.add_parser("review_pack")
    rp.add_argument("--data_dir", required=True)
    rp.add_argument("--out_dir", required=True)
    rp.add_argument("--n_extractions", type=int, default=120)
    rp.add_argument("--n_edges", type=int, default=120)
    rp.add_argument("--strategy", default="mixed", choices=["mixed","random"])

    lj = sub.add_parser("llm_judge")
    lj.add_argument("--data_dir", required=True)
    lj.add_argument("--model", required=True, help="HF model name, e.g., Qwen/Qwen2.5-7B-Instruct")
    lj.add_argument("--device", default="auto", choices=["auto","cpu"])
    lj.add_argument("--max_new_tokens", type=int, default=96)
    lj.add_argument("--temperature", type=float, default=0.0)
    lj.add_argument("--judge_edges", default="true")
    lj.add_argument("--judge_extractions", default="true")
    lj.add_argument("--out_file", default="")
    lj.add_argument("--cache_file", default="")

    st = sub.add_parser("stats")
    st.add_argument("--data_dir", required=True)
    st.add_argument("--judge_file", required=True)
    st.add_argument("--human_csv", default="", help="Optional filled human review CSV to compute kappa")

    args = ap.parse_args()

    if args.cmd == "build":
        cmd_build(args)
    elif args.cmd == "validatepp":
        cmd_validatepp(args)
    elif args.cmd == "review_pack":
        cmd_review_pack(args)
    elif args.cmd == "llm_judge":
        cmd_llm_judge(args)
    elif args.cmd == "stats":
        cmd_stats(args)

if __name__ == "__main__":
    main()
